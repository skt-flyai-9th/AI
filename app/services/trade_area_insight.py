"""가게 상권분석 인사이트 생성 (마이페이지 "내 가게 상권 분석" 화면용).

`store`(이름/카테고리/주소/좌표)를 받아 `상권분석DB.xlsx`에 실제로 있는 값만 근거로
① 상권 이름 ② 한 줄 요약(LLM) ③ 나이대·성별 분포(%)를 만든다.

**알려진 한계** — 공식 상권 폴리곤 경계 데이터가 아직 없어(`data_quality_log`
QA-FINAL-015 FAIL 참고), 위치 매칭은 상권 대표 좌표까지의 직선거리로 가장 가까운
곳을 고르는 근사치다. 실제 폴리곤이 확보되면 `_match_by_coordinates`만 point-in-
polygon으로 교체하면 된다.
"""

from __future__ import annotations

import json
import logging
import math
from dataclasses import dataclass
from functools import lru_cache
from importlib import resources
from typing import Any

import httpx
import openpyxl

from app.core.config import get_settings
from app.schemas.trade_area_insight import (
    AgeDistribution,
    GenderDistribution,
    StoreInfo,
    TradeAreaInsightResponse,
)

logger = logging.getLogger(__name__)

_SOURCE_PACKAGE = "app.template_knowledge.sources"
_WORKBOOK_NAME = "상권분석DB.xlsx"

# 폴리곤이 없어 대표 좌표 간 직선거리로 근사 매칭한다. 이 거리(m)를 넘으면
# "이 근처에 뚜렷한 공식 상권이 없다"로 보고 district_name을 null로 돌려준다.
_MAX_MATCH_DISTANCE_M = 500.0

_AGE_KEYS = ("age10", "age20", "age30", "age40", "age50", "age60_plus")


@dataclass(frozen=True)
class OfficialTradeArea:
    trdar_cd: str
    name: str
    district_name: str | None
    admin_dong_name: str | None
    latitude: float | None
    longitude: float | None


@dataclass(frozen=True)
class RegionProfile:
    region_id: str
    name: str
    area_type: str | None
    main_visitors: str | None
    tone: str | None
    trend_notes: str | None
    nearby_landmarks: str | None


@dataclass(frozen=True)
class FloatPopulationFact:
    total: float
    male: float
    female: float
    age: dict[str, float]


@dataclass(frozen=True)
class SalesFact:
    male_count: float
    female_count: float
    age_count: dict[str, float]


@dataclass(frozen=True)
class TradeAreaKnowledge:
    areas: list[OfficialTradeArea]
    area_to_region_id: dict[str, str]
    regions_by_id: dict[str, RegionProfile]
    category_name_to_sarils_id: dict[str, str]
    sarils_id_to_service_codes: dict[str, list[str]]
    latest_float_period: str | None
    float_by_trdar: dict[str, FloatPopulationFact]
    citywide_float: FloatPopulationFact | None
    latest_sales_period: str | None
    sales_by_trdar_and_service: dict[tuple[str, str], SalesFact]


def _open_workbook():
    package = resources.files(_SOURCE_PACKAGE)
    with resources.as_file(package.joinpath(_WORKBOOK_NAME)) as path:
        return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _header_index(ws) -> dict[str, int]:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(name): idx for idx, name in enumerate(header) if name is not None}


def _f(value: Any) -> float:
    try:
        return float(value) if value is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


@lru_cache(maxsize=1)
def load_trade_area_knowledge() -> TradeAreaKnowledge:
    wb = _open_workbook()
    try:
        areas, area_to_region_id, regions_by_id = _load_geography(wb)
        category_name_to_sarils_id, sarils_id_to_service_codes = _load_category_maps(wb)
        latest_float_period, float_by_trdar, citywide_float = _load_float_population(wb)
        latest_sales_period, sales_by_trdar_and_service = _load_sales(wb)
    finally:
        wb.close()
    return TradeAreaKnowledge(
        areas=areas,
        area_to_region_id=area_to_region_id,
        regions_by_id=regions_by_id,
        category_name_to_sarils_id=category_name_to_sarils_id,
        sarils_id_to_service_codes=sarils_id_to_service_codes,
        latest_float_period=latest_float_period,
        float_by_trdar=float_by_trdar,
        citywide_float=citywide_float,
        latest_sales_period=latest_sales_period,
        sales_by_trdar_and_service=sales_by_trdar_and_service,
    )


def _load_geography(wb):
    ws = wb["official_trade_areas"]
    idx = _header_index(ws)
    areas: list[OfficialTradeArea] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        trdar_cd = row[idx["current_trdar_cd"]]
        if trdar_cd is None:
            continue
        lat = row[idx["latitude_wgs84_2023"]]
        lon = row[idx["longitude_wgs84_2023"]]
        areas.append(
            OfficialTradeArea(
                trdar_cd=str(trdar_cd),
                name=str(row[idx["official_name_2023"]] or ""),
                district_name=row[idx["district_name"]],
                admin_dong_name=row[idx["admin_dong_name_current"]],
                latitude=float(lat) if lat is not None else None,
                longitude=float(lon) if lon is not None else None,
            )
        )

    ws = wb["content_region_trade_area_map"]
    idx = _header_index(ws)
    area_to_region_id: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        trdar_cd = row[idx["current_trdar_cd"]]
        region_id = row[idx["content_region_id"]]
        status = row[idx["status"]]
        if trdar_cd is None or region_id is None:
            continue
        if str(status).lower() != "approved":
            continue
        area_to_region_id[str(trdar_cd)] = str(region_id)

    ws = wb["regions"]
    idx = _header_index(ws)
    regions_by_id: dict[str, RegionProfile] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        region_id = row[idx["region_id"]]
        if region_id is None:
            continue
        regions_by_id[str(region_id)] = RegionProfile(
            region_id=str(region_id),
            name=str(row[idx["name"]] or ""),
            area_type=row[idx.get("area_type", -1)] if "area_type" in idx else None,
            main_visitors=row[idx["main_visitors(주 방문층)"]] if "main_visitors(주 방문층)" in idx else None,
            tone=row[idx["tone(상권 톤 한 줄)"]] if "tone(상권 톤 한 줄)" in idx else None,
            trend_notes=row[idx["trend_notes(최근 변화)"]] if "trend_notes(최근 변화)" in idx else None,
            nearby_landmarks=row[idx["nearby_landmarks(랜드마크/역)"]] if "nearby_landmarks(랜드마크/역)" in idx else None,
        )
    return areas, area_to_region_id, regions_by_id


def _load_category_maps(wb):
    ws = wb["categories"]
    idx = _header_index(ws)
    category_name_to_sarils_id: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        category_id = row[idx["category_id"]]
        name = row[idx["name"]]
        group = row[idx.get("category_group", -1)] if "category_group" in idx else None
        if category_id is None:
            continue
        if name:
            category_name_to_sarils_id[str(name)] = str(category_id)
        if group:
            category_name_to_sarils_id.setdefault(str(group), str(category_id))

    ws = wb["official_category_map"]
    idx = _header_index(ws)
    sarils_id_to_service_codes: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[idx["official_service_code"]]
        sarils_id = row[idx["sarils_primary_category_id"]]
        if code is None or sarils_id is None:
            continue
        sarils_id_to_service_codes.setdefault(str(sarils_id), []).append(str(code))
    return category_name_to_sarils_id, sarils_id_to_service_codes


def _load_float_population(wb):
    ws = wb["fact_float_population_qtr"]
    idx = _header_index(ws)
    latest_period: str | None = None
    rows_by_trdar: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        trdar_cd = row[idx["trdar_cd"]]
        period = row[idx["period"]]
        if trdar_cd is None or period is None:
            continue
        period = str(period)
        if latest_period is None or period > latest_period:
            latest_period = period
        rows_by_trdar[str(trdar_cd)] = {
            "period": period,
            "total": _f(row[idx["total_float"]]),
            "male": _f(row[idx["male_float"]]),
            "female": _f(row[idx["female_float"]]),
            "age10": _f(row[idx["age10_float"]]),
            "age20": _f(row[idx["age20_float"]]),
            "age30": _f(row[idx["age30_float"]]),
            "age40": _f(row[idx["age40_float"]]),
            "age50": _f(row[idx["age50_float"]]),
            "age60_plus": _f(row[idx["age60_plus_float"]]),
        }

    float_by_trdar: dict[str, FloatPopulationFact] = {}
    totals = {"total": 0.0, "male": 0.0, "female": 0.0, **{k: 0.0 for k in _AGE_KEYS}}
    for trdar_cd, values in rows_by_trdar.items():
        if latest_period is not None and values["period"] != latest_period:
            continue
        fact = FloatPopulationFact(
            total=values["total"],
            male=values["male"],
            female=values["female"],
            age={k: values[k] for k in _AGE_KEYS},
        )
        float_by_trdar[trdar_cd] = fact
        totals["male"] += fact.male
        totals["female"] += fact.female
        for k in _AGE_KEYS:
            totals[k] += fact.age[k]

    citywide = (
        FloatPopulationFact(
            total=totals["male"] + totals["female"],
            male=totals["male"],
            female=totals["female"],
            age={k: totals[k] for k in _AGE_KEYS},
        )
        if float_by_trdar
        else None
    )
    return latest_period, float_by_trdar, citywide


def _load_sales(wb):
    ws = wb["fact_sales_qtr"]
    idx = _header_index(ws)
    latest_period: str | None = None
    raw: dict[tuple[str, str], dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        trdar_cd = row[idx["trdar_cd"]]
        svc = row[idx["svc_industry_cd"]]
        period = row[idx["period"]]
        if trdar_cd is None or svc is None or period is None:
            continue
        period = str(period)
        if latest_period is None or period > latest_period:
            latest_period = period
        raw[(str(trdar_cd), str(svc))] = {
            "period": period,
            "male_cnt": _f(row[idx["male_sales_count"]]),
            "female_cnt": _f(row[idx["female_sales_count"]]),
            "age10_cnt": _f(row[idx["age10_cnt"]]),
            "age20_cnt": _f(row[idx["age20_cnt"]]),
            "age30_cnt": _f(row[idx["age30_cnt"]]),
            "age40_cnt": _f(row[idx["age40_cnt"]]),
            "age50_cnt": _f(row[idx["age50_cnt"]]),
            "age60_plus_cnt": _f(row[idx["age60_plus_cnt"]]),
        }

    sales_by_key: dict[tuple[str, str], SalesFact] = {}
    for key, values in raw.items():
        if latest_period is not None and values["period"] != latest_period:
            continue
        sales_by_key[key] = SalesFact(
            male_count=values["male_cnt"],
            female_count=values["female_cnt"],
            age_count={
                "age10": values["age10_cnt"],
                "age20": values["age20_cnt"],
                "age30": values["age30_cnt"],
                "age40": values["age40_cnt"],
                "age50": values["age50_cnt"],
                "age60_plus": values["age60_plus_cnt"],
            },
        )
    return latest_period, sales_by_key


def _haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def _match_by_coordinates(
    knowledge: TradeAreaKnowledge, latitude: float, longitude: float
) -> OfficialTradeArea | None:
    best: OfficialTradeArea | None = None
    best_distance = math.inf
    for area in knowledge.areas:
        if area.latitude is None or area.longitude is None:
            continue
        distance = _haversine_m(latitude, longitude, area.latitude, area.longitude)
        if distance < best_distance:
            best_distance = distance
            best = area
    if best is None or best_distance > _MAX_MATCH_DISTANCE_M:
        return None
    return best


def _match_by_address(knowledge: TradeAreaKnowledge, address: str) -> OfficialTradeArea | None:
    """폴리곤도 좌표도 없을 때의 최후 수단 — 주소 문자열에 행정동/상권명이 그대로
    들어있는 경우만 잡아낸다. 정밀 지오코딩이 아니므로 놓치는 경우가 많을 수 있다."""
    normalized = address.replace(" ", "")
    candidates = [
        area
        for area in knowledge.areas
        if area.admin_dong_name and str(area.admin_dong_name).replace(" ", "") in normalized
    ]
    if not candidates:
        return None
    # 여러 상권이 같은 동에 걸치면 상권명이 주소에 직접 등장하는 쪽을 우선한다.
    for area in candidates:
        if area.name and area.name.replace(" ", "") in normalized:
            return area
    return candidates[0]


def _resolve_service_codes(knowledge: TradeAreaKnowledge, store: StoreInfo) -> list[str]:
    for text in (store.sub_category, store.category):
        if not text:
            continue
        sarils_id = knowledge.category_name_to_sarils_id.get(text)
        if sarils_id is None:
            for name, candidate_id in knowledge.category_name_to_sarils_id.items():
                if name in text or text in name:
                    sarils_id = candidate_id
                    break
        if sarils_id:
            codes = knowledge.sarils_id_to_service_codes.get(sarils_id)
            if codes:
                return codes
    return []


def _normalize_to_100(values: dict[str, float]) -> dict[str, int]:
    """가장 큰 값에 나머지를 몰아줘서 합이 항상 정확히 100이 되게 한다."""
    total = sum(values.values())
    if total <= 0:
        # 데이터가 전혀 없으면 균등 분배(합계 100 보장이 최우선 계약이라 임의 추정보다 안전).
        n = len(values)
        base, remainder = divmod(100, n)
        keys = list(values.keys())
        return {k: base + (1 if i < remainder else 0) for i, k in enumerate(keys)}
    raw = {k: (v / total) * 100 for k, v in values.items()}
    rounded = {k: int(v) for k, v in raw.items()}
    remainder = 100 - sum(rounded.values())
    if remainder:
        order = sorted(raw, key=lambda k: raw[k] - rounded[k], reverse=True)
        for k in order[:remainder]:
            rounded[k] += 1
    return rounded


def _compute_distributions(
    knowledge: TradeAreaKnowledge,
    matched_area: OfficialTradeArea | None,
    store: StoreInfo,
) -> tuple[AgeDistribution, GenderDistribution]:
    age_raw: dict[str, float] | None = None
    male_raw = 0.0
    female_raw = 0.0

    if matched_area is not None:
        service_codes = _resolve_service_codes(knowledge, store)
        sales_fact = next(
            (
                knowledge.sales_by_trdar_and_service[(matched_area.trdar_cd, code)]
                for code in service_codes
                if (matched_area.trdar_cd, code) in knowledge.sales_by_trdar_and_service
            ),
            None,
        )
        if sales_fact is not None:
            age_raw = dict(sales_fact.age_count)
            male_raw, female_raw = sales_fact.male_count, sales_fact.female_count
        else:
            float_fact = knowledge.float_by_trdar.get(matched_area.trdar_cd)
            if float_fact is not None:
                age_raw = dict(float_fact.age)
                male_raw, female_raw = float_fact.male, float_fact.female

    if age_raw is None and knowledge.citywide_float is not None:
        age_raw = dict(knowledge.citywide_float.age)
        male_raw, female_raw = knowledge.citywide_float.male, knowledge.citywide_float.female

    if age_raw is None:
        age_raw = {k: 1.0 for k in _AGE_KEYS}
        male_raw = female_raw = 1.0

    age_5bucket = {
        "10s": age_raw["age10"],
        "20s": age_raw["age20"],
        "30s": age_raw["age30"],
        "40s": age_raw["age40"],
        "50s_plus": age_raw["age50"] + age_raw["age60_plus"],
    }
    age_pct = _normalize_to_100(age_5bucket)
    gender_pct = _normalize_to_100({"male": male_raw, "female": female_raw})
    return (
        AgeDistribution.model_validate(age_pct),
        GenderDistribution.model_validate(gender_pct),
    )


class TradeAreaSummaryError(RuntimeError):
    pass


def _generate_summary(region: RegionProfile | None, store: StoreInfo) -> str:
    settings = get_settings()
    subject = store.sub_category or store.category
    if region is None:
        return f"{store.name}({subject}) 주변은 뚜렷한 대표 상권으로 특정하기 어려운 생활권입니다."

    if not settings.openai_api_key.strip():
        # 연동 전(로컬/테스트)에는 규칙 기반 문장으로 대체한다 — 지어내지 않고 DB에
        # 실제로 있는 값만 이어붙인다.
        visitors = region.main_visitors or "다양한 방문객"
        return f"{visitors}이 많이 찾는 {region.name} 상권으로, {subject} 업종과 어울리는 상권입니다."

    payload = {
        "region_name": region.name,
        "area_type": region.area_type,
        "main_visitors": region.main_visitors,
        "tone": region.tone,
        "trend_notes": region.trend_notes,
        "nearby_landmarks": region.nearby_landmarks,
        "store_category": store.category,
        "store_sub_category": store.sub_category,
    }
    request_payload = {
        "model": settings.database_openai_model,
        "instructions": (
            "당신은 사릴스 상권분석 요약 작성자입니다. 주어진 상권 프로필과 가게 업종 정보만 "
            "근거로 한국어 한 문장을 만드세요. 프로필에 없는 사실(구체적 수치, 매출, 사건 등)을 "
            "지어내지 마세요. 화면 카드 한 줄에 들어가므로 간결해야 합니다."
        ),
        "input": [
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": json.dumps(payload, ensure_ascii=False)}
                ],
            }
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "trade_area_summary",
                "schema": {
                    "type": "object",
                    "properties": {"summary": {"type": "string"}},
                    "required": ["summary"],
                    "additionalProperties": False,
                },
                "strict": True,
            }
        },
        "max_output_tokens": 300,
        "store": False,
    }
    try:
        with httpx.Client(timeout=settings.database_request_timeout_seconds) as client:
            response = client.post(
                f"{settings.openai_base_url.rstrip('/')}/responses",
                headers={
                    "Authorization": f"Bearer {settings.openai_api_key.strip()}",
                    "Content-Type": "application/json",
                },
                json=request_payload,
            )
        response.raise_for_status()
        data = response.json()
        text = None
        for item in data.get("output", []):
            for part in item.get("content", []):
                if part.get("type") == "output_text":
                    text = part.get("text")
        if text is None:
            text = data.get("output_text")
        parsed = json.loads(text)
        summary = str(parsed["summary"]).strip()
        if not summary:
            raise ValueError("empty summary")
        return summary
    except Exception as exc:  # noqa: BLE001 - LLM 실패는 규칙 기반으로 안전하게 대체
        logger.warning("상권 요약 LLM 생성 실패, 규칙 기반으로 대체: %s", exc)
        visitors = region.main_visitors or "다양한 방문객"
        return f"{visitors}이 많이 찾는 {region.name} 상권으로, {subject} 업종과 어울리는 상권입니다."


def build_trade_area_insight(store: StoreInfo) -> TradeAreaInsightResponse:
    knowledge = load_trade_area_knowledge()

    matched_area: OfficialTradeArea | None = None
    if store.latitude is not None and store.longitude is not None:
        matched_area = _match_by_coordinates(knowledge, store.latitude, store.longitude)
    if matched_area is None:
        matched_area = _match_by_address(knowledge, store.address)

    region: RegionProfile | None = None
    if matched_area is not None:
        region_id = knowledge.area_to_region_id.get(matched_area.trdar_cd)
        if region_id is not None:
            region = knowledge.regions_by_id.get(region_id)

    age_distribution, gender_distribution = _compute_distributions(knowledge, matched_area, store)
    summary = _generate_summary(region, store)

    return TradeAreaInsightResponse(
        district_name=region.name if region is not None else None,
        summary=summary,
        age_distribution=age_distribution,
        gender_distribution=gender_distribution,
    )
